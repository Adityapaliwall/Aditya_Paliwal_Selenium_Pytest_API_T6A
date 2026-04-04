import pytest
import openpyxl
from openpyxl.workbook import Workbook

wb = openpyxl.Workbook()
sheetName = "sheet1"
if sheetName in wb.sheetnames:
    ws = wb[sheetName]
else:
    ws = wb.create_sheet(sheetName)

ws['A1'] = 'user'
ws['B1'] = 'password'

ws.append(['user1', '123'])
ws.append(['user2', '456'])
ws.append(['user3', '789'])
ws.append(['user4', '789'])

ws['A8'] = 'use8'
ws['B8'] = '987'
ws.append(['user9', '789'])

ws.delete_rows(3)

for row in ws.iter_rows(values_only=True):              ## extract the dataa from the excel sheet through in build method
    print(row)

wb.save('sample.xlsx')
### wb.save(r"C:\Users\ADITYA\Downloads\samplee.xlsx")          save in the given path